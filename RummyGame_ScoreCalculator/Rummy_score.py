import openpyxl,os
from datetime import date
import pandas as pd
##presetup
os.system('cls')
print(" **********  Maintaining single sheet per day ***********")

def  GetDate():
    '''Function to get the current date'''
    today=date.today()
    datedetails=today.strftime("%d-%m-%y")
    return datedetails
sheetName=GetDate()

## Check for the sheet with same date ###
xlsheet="Rummy_score.xlsx"

def CreateFile(inputFile):
    if os.path.isfile(inputFile):
        print(">>>>>> Already Excel File Found")
        return True
    else:
        print("New File Created <<<<<<<")
        writer = pd.ExcelWriter(inputFile, engine='xlsxwriter')
        writer.save()
    return True
print(CreateFile(xlsheet))

def CheckSheetFound(xlfile,sheet):
    wb = openpyxl.load_workbook(xlfile)
    try:
        wb[str(sheet)]
        print(">>>>>> Already Excel sheet Found")
    except:
        ws1=wb.create_sheet(str(sheet))

        print("Created sheet <<<<<<<")
        sheets = wb.sheetnames
        import re
        for i in sheets:
            if re.match("Sheet",i):
                wb.remove(wb[i])
                print("Removed extra sheets")
        wb.save(xlfile)
CheckSheetFound(xlsheet,sheetName)

##GetNumber of users
numberOfusers=int(input("Please Enter Number of users \n"))
Userlist=[]
for i in range(numberOfusers):
    k=input("Enter User name \n")
    Userlist.append(str(k))
Users={}
Users={k:[0] for k in Userlist}

def getPoints(Users):
    for j in range(1,8):
        for i in Users.keys():
            points=int(input(" Enter your points for the round "+str(j)+" For the user "+str(i))+"\n")
            Users[i].append(points)
    return Users
Results=getPoints(Users)

for h in Users:
    Users[h].append(sum(Users[h]))
df=pd.DataFrame(Results)
df.to_excel(xlsheet,sheet_name=sheetName)

dummy={}
dummy={k:[0] for k in Userlist}
for i in dummy.keys():
    dummy[i]=Users[i][8]
Results={k: v for k, v in sorted(dummy.items(), key=lambda item: item[1],reverse=False)}
for c,k in enumerate(Results):
    print(str(k)+" in position "+str(int(c)+1))
print("The winner is "+str(list(Results.keys())[0]))