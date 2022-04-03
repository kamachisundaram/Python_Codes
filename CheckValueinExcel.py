import openpyxl,os
from datetime import date
import pandas as pd
##presetup
os.system('cls')
print(" **********  Maintaining single sheet per day ***********")

xlsheet="TestFile.xlsx"
## Check for the sheet with same date ###

if os.path.isfile(xlsheet):
    print(">>>>>> Already Excel File Found")

else:
    print("Data missing")
    raise DataError
wb_obj = openpyxl.load_workbook(xlsheet)
sheet_obj = wb_obj.active

dummy={k:[0] for k in range(sheet_obj.max_row)}
for i in range(sheet_obj.max_row):
    dummy[i]=sheet_obj.cell(row=i,column=1)



'''
import pandas as pd
df = pd.read_excel (xlsheet)
df = pd.DataFrame(df, columns= ['Score'])
for index,row in df.iterrows():
    print(row)

print(df)
'''
def CheckSheetFound(xlfile,sheet):
    wb = openpyxl.load_workbook(xlfile)
    try:
        wb[str(sheet)]
        print(">>>>>> Already Excel sheet Found")
    except:
        ws1=wb.create_sheet(str(sheet))

        print("Created sheet <<<<<<<")
        sheets = wb.sheetnames
        import re
        for i in sheets:
            if re.match("Sheet",i):
                wb.remove(wb[i])
                print("Removed extra sheets")
        wb.save(xlfile)



print(dummy)
